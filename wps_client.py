# -*- coding: utf-8 -*-
"""WPS 云盘 API 客户端 — 支持本地 cookies.xlsx 和 Streamlit Cloud secrets"""
import urllib.request, ssl, json, io
ssl._create_default_https_context = ssl._create_unverified_context

GROUP_ID = 3055417161

def _load_cookies():
    """加载认证凭据：优先 Streamlit secrets，其次本地 cookies.xlsx"""
    import os
    try:
        import streamlit as st
        required = ['WPS_SID', 'KSO_SID', 'WPS_CSRF', 'WPS_UID']
        missing = [k for k in required if k not in st.secrets or not st.secrets[k]]
        if not missing:
            return {
                'wps_sid': st.secrets['WPS_SID'],
                'kso_sid': st.secrets['KSO_SID'],
                'csrf': st.secrets['WPS_CSRF'],
                'uid': st.secrets['WPS_UID'],
                'cv': st.secrets.get('WPS_CV', ''),
                'env': st.secrets.get('WPS_ENV', 'prod-0'),
            }
        if missing:
            raise KeyError(f'Streamlit Cloud Secrets 缺失或为空: {", ".join(missing)}')
    except KeyError:
        raise
    except Exception:
        pass

    # Fallback: 本地 cookies.xlsx
    if not os.path.exists('cookies.xlsx'):
        raise FileNotFoundError(
            'Streamlit Cloud Secrets 未配置，且本地 cookies.xlsx 不存在。'
            '请在 Streamlit Cloud → Settings → Secrets 中添加: '
            'WPS_SID / KSO_SID / WPS_CSRF / WPS_UID / WPS_CV / WPS_ENV'
        )
    import openpyxl
    wb = openpyxl.load_workbook('cookies.xlsx')
    ws = wb['Sheet2']
    cookies = {}
    for row in ws.iter_rows(values_only=True):
        name = row[0]
        value = row[1]
        domain = str(row[2]) if row[2] else ''
        if name and value and '.kdocs.cn' in domain:
            if name not in cookies:
                cookies[name] = value
    return cookies

COOKIES = _load_cookies()
COOKIE_STR = '; '.join(f'{k}={v}' for k, v in COOKIES.items())
CSRF = COOKIES.get('csrf', '')

def _headers():
    return {
        'Cookie': COOKIE_STR,
        'x-csrf-rand': CSRF,
        'Accept': 'application/json',
    }

def list_files(parentid=0):
    url = (f'https://drive.wps.cn/api/v5/groups/{GROUP_ID}/files'
           f'?count=200&offset=0&order=asc&orderby=fname'
           f'&parentid={parentid}&with_link=true')
    req = urllib.request.Request(url, headers=_headers())
    r = urllib.request.urlopen(req, timeout=15)
    return json.loads(r.read()).get('files', [])

def download_file(file_id):
    url = (f'https://drive.wps.cn/api/v5/groups/{GROUP_ID}/files/{file_id}/download'
           f'?isblocks=0&support_checksums=md5,sha1')
    req = urllib.request.Request(url, headers=_headers())
    r = urllib.request.urlopen(req, timeout=15)
    presigned_url = json.loads(r.read()).get('url', '')
    if not presigned_url:
        raise Exception('无法获取下载链接')
    req2 = urllib.request.Request(presigned_url)
    r2 = urllib.request.urlopen(req2, timeout=60)
    return r2.read()

def read_xlsx(file_id):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    orig_init = DataValidation.__init__
    def patched(self, *a, **kw):
        for k in ['id', 'uid', 'xr', 'sqref', 'extLst']:
            kw.pop(k, None)
        return orig_init(self, *a, **kw)
    DataValidation.__init__ = patched
    content = download_file(file_id)
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True)
